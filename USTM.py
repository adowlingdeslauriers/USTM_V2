''' TODO
-Make sure there are no FDA clients in S321
-Better handling for duplicate orders
	-USTM assumes 1 order number => 1 package. This may no longer be true
-How to identify if only some orders in a batch are matched to the ACE manifest
	-eg. LUS
	-Add 3PLC fetches? Get order lists for batches?
-SCNs Duplicate Remover
-Stop it from crashing hard when there are gaylord conflicts

-test

-Add version designators for ACEs
-blacklist tvidler
-remake exe
'''

import json
from datetime import date
from datetime import datetime
import datetime as datetime_module
import os
import csv
import traceback
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from email.mime.base import MIMEBase
from email import encoders
import requests
import math

#installable through PiP
import openpyxl as pyxl
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.utils import Image

#GUI library, included in /appJar folder
from appJar import gui

### CLasses

class Config():
	data = None
	def __init__(self):
		self.load()

	def load(self):
		with open("resources/CONFIG.json", "r") as file:
			self.data = json.load(file)

	def save(self):
		with open("resources/CONFIG.json", "w") as file:
			json.dump(self.data, file, indent = 4)

#Empty trip, for storing trip data
class Trip():
	date = None
	BoL = None
	PAPS = None
	SCAC = None
	gaylord_assignment = None
	gaylord_count = None
	usps_count = None
	dhl_count = None
	fedex_count = None
	package_count = None
	total_weight = None
	shipper = None

### Functions

## Mainline Functions

def init_UI():
	#Tabbed Frame
	app.startTabbedFrame("TabbedFrame")
	app.setSticky("nesw")
	app.setStretch("column")

	## Frame 1
	app.startTab("BASIC")

	app.startLabelFrame("Date/BoL/PAPS")
	app.addLabel("W:\\Logistics\\Carrier Tracking\\USPS Tracking.xlsx")
	app.addLabelEntry("Date:")
	app.addLabelEntry("BoL #:")
	app.addLabelEntry("PAPS #:")
	app.addLabelOptionBox("Shipper:", [shipper_profiles for shipper_profiles in config.data["shipper_profiles"]])
	app.addButton("+1 BoL #/PAPS #", increase_trip_variables)
	app.addButton("-1 BoL #/PAPS #", decrease_trip_variables)
	app.addButton("Save BoL #/PAPS #", save_trip_variables)
	app.stopLabelFrame()

	app.startLabelFrame("Step 1: Creating Trip & ACE")
	app.addLabel("ACE Manfiest (from Techship, \"ace_manifest_#\"):")
	app.addFileEntry("ACEManifestFileEntry")
	app.addLabel("Report XLSX (from Techship, \"manifest_packages_#\"):")
	app.addFileEntry("XLSXReportFileEntry")
	app.addLabel("USTM Scans File:")
	app.addFileEntry("batchesFileEntry")
	app.addButton("Create Paperwork", create_paperwork)
	app.stopLabelFrame()

	app.startLabelFrame("Step 2: Emailing Trip")
	app.addLabel("trip_summary_label", "No Trip Summary found! Can not email")
	app.addFileEntry("Trip.info file:")
	app.addButton("Load", load_trip_summary)

	app.addLabel("ProForma (Printed from SmartBorder):")
	app.addFileEntry("ProFormaFileEntry")
	app.addLabelEntry("Email Subject:")
	app.addLabel("Email Text:")

	app.setStretch("both")
	app.setSticky("nesw")
	app.addTextArea("EmailTextArea", text = config.data["default_email_message"])

	app.setSticky("ws")
	app.setStretch("column")
	app.addButton("Email Paperwork", email_paperwork)
	app.stopLabelFrame()
	app.stopTab()

	## Frame 2
	app.startTab("ACE EDITING")

	app.startLabelFrame("ACE Manifest")
	app.addLabel("ACE Manifest (.json):")
	app.addFileEntry("ACEManifestFileEntry2")
	app.addLabelEntry("File Date:")
	app.addButton("Load ACE", load_ACE_manifest)
	app.addLabel("ACE_status_label", "No ACE Loaded")
	app.stopLabelFrame()

	app.startLabelFrame("Gaylord Removal")
	app.addLabelEntry('Gaylord (eg. "G1"):')
	app.addButton("Remove Gaylord", remove_gaylord)
	app.stopLabelFrame()

	app.startLabelFrame("Batch/Order Removal")
	app.addLabel('Batches/Orders/SCACs:')
	app.addTextArea("batchesTextArea")
	app.addButton("Remove Items", remove_items)
	app.stopLabelFrame()

	app.startLabelFrame("Duplicate SCN Editor")
	app.addLabel("SCNLabel", "SCNs currently end with: NA")
	app.addLabelEntry("New 2 digits:")
	app.addButton("Change SCNs", change_SCNs)
	app.stopLabelFrame()

	app.startLabelFrame("ACE Splitter")
	app.addLabel("Use if ACE exceeds 9999 entries:")
	app.addButton("Split", split_ACE)
	app.stopLabelFrame()

	app.stopTab()

	## Frame 3
	app.startTab("ADVANCED")

	app.startLabelFrame("Convert ACE to EXCEL:")
	app.addLabelFileEntry("JSON")
	app.addButton("Convert to CSV", convert_JSON_to_CSV)
	app.addLabelFileEntry("CSV")
	app.addButton("Convert to JSON", convert_CSV_to_JSON)
	app.stopLabelFrame()

	app.startLabelFrame("JSON Formatter")
	app.addLabel("Overwrites current JSON")
	app.addLabelFileEntry("Ugly JSON")
	app.addButton("Format JSON", json_beautifier)
	app.stopLabelFrame()

	app.startLabelFrame("JSON Combiner")
	app.addLabel("Outputs to today's folder")
	app.addLabelFileEntry("JSON 1")
	app.addLabelFileEntry("JSON 2")
	app.addButton("Combine", combine_JSONs)
	app.stopLabelFrame()

	app.startLabelFrame("Warnable Clients")
	app.addLabel('Clients:')
	app.addTextArea("warnableClientsTextArea")
	app.addButton("Save List", save_warnable_clients)
	app.stopLabelFrame()

	app.stopTab()

	## Frame 4
	app.startTab("USGR")
	app.startLabelFrame("USGR")
	app.addLabelFileEntry("USGR Data:")
	app.addLabelEntry("USGR Date:")
	app.addLabelEntry("USGR Entry Number:")
	app.addLabelEntry("USGR BoL #:")
	app.addButton("Create USGR", create_USGR)
	app.stopLabelFrame()
	app.stopTab()

	## Finish GUI
	app.stopTabbedFrame()
	load_variables()
	app.go()

def load_variables():
	#CONFIG.json is used to keep track of what the latest BoL/PAPS number is
	
	config.data["default_date"] = str(date.today())
	app.setEntry("Date:", config.data["default_date"])
	app.setEntry("BoL #:", config.data["default_BoL"])
	app.setEntry("PAPS #:", config.data["default_PAPS"])
	app.setEntry("USGR Date:", config.data["default_date"])
	app.setEntry("File Date:", config.data["default_date"])
	app.setEntry("Email Subject:", config.data["default_email_subject"])

	output = ""
	for line in config.data["warnable_clients"]:
		output += line + "\n"
	app.setTextArea("warnableClientsTextArea", output)
	
def increase_trip_variables():
	#Sets date to today and advances BoL and PAPS number by 1
	config.data["default_date"] = str(date.today())
	app.setEntry("Date:", config.data["default_date"])

	config.data["default_BoL"] = str(int(config.data["default_BoL"]) + 1).zfill(7)
	app.setEntry("BoL #:", config.data["default_BoL"])

	config.data["default_PAPS"] = str(int(config.data["default_PAPS"]) + 1).zfill(6)
	app.setEntry("PAPS #:", config.data["default_PAPS"])

def decrease_trip_variables():
	config.data["default_date"] = str(date.today())
	app.setEntry("Date:", config.data["default_date"])

	config.data["default_BoL"] = str(int(config.data["default_BoL"]) - 1).zfill(7)
	app.setEntry("BoL #:", config.data["default_BoL"])

	config.data["default_PAPS"] = str(int(config.data["default_PAPS"]) - 1).zfill(6)
	app.setEntry("PAPS #:", config.data["default_PAPS"])

def save_trip_variables():
	config.data["default_date"] = app.getEntry("Date:")
	config.data["default_BoL"] = app.getEntry("BoL #:")
	config.data["default_PAPS"] = app.getEntry("PAPS #:")
	with open("resources/CONFIG.json", "w") as variables_file:
		json.dump(config.data, variables_file, indent = 4)

def save_warnable_clients():
	text = app.getTextArea("warnableClientsTextArea")
	lines = [line for line in text.split("\n") if line != ""]
	config.data["warnable_clients"] = lines
	config.save()

def create_paperwork():
	"""Does everything required to make the paperwork for the day.
	Draws a lot from Config and Trip
	"""
	trip.date = app.getEntry("Date:")
	trip.BoL = app.getEntry("BoL #:")
	trip.PAPS = app.getEntry("PAPS #:")
	print(f"\nCreating trip with BoL {trip.BoL} and PAPS {trip.PAPS}")

	for shipper in config.data["shipper_profiles"]:
		if shipper == app.getOptionBox("Shipper:"):
			trip.shipper = config.data["shipper_profiles"][shipper]
			print("Shipper set to:")
			print(trip)

	create_output_folder(trip.date)
	consolidated_json = create_consolidated_JSON()
	master_ACE, S321_ACE, good_orders = construct_ACE(consolidated_json) #Merges XLSX report (carrier, client, tracking number, close date) with ACE report (shipping info) and ACE CSVR Report (SKU names) to create one unified data source
																				#Then filters out all ACE entries that don't match a batch scan. Also outputs ACE
	assign_gaylords(master_ACE) #Creates a list of all gaylords, which carrier (eg. DHL, USPS, FedEx) they belong to, if it has FDA-regulated products

	#Uses the core data to produce paperwork 
	create_detailed_report(master_ACE) #The report Tri-Ad uses to find packages when searched
	create_loadsheet() #Creates a useful but non-necessary info sheet
	create_BoL() #Creates the Bill of lading for this load
	create_IMS_BoL() #Creates the Bill of Lading for IMS
	create_ProForma(master_ACE) #Creates the XLSX file we upload to SmartBorder to create our Proforma invoice
	create_detailed_report_CSV(S321_ACE)
	create_trip_summary()
	#createBorderConnectCSV() #Creates a .csv version of the ACE JSON for BorderConnect

	message_box("Success!\nPaperwork generated")

def create_consolidated_JSON():
	"""the ACE, ACE CSV and XLSX report. Matches entries in the two based on order IDs. Takes the client/carrier/close date data from the XLSX and adds it to the ACE."""

	#Loads unprocessed ACE manifest
	ACE_data = load_ACE()
	XLSX_data, ORDERID_column_index, client_name_column_index, carrier_column_index, ship_date_column_index, tracking_number_column_index = load_manifest_report()

	#Matches XLSX Report data to ACE Manifest data using ORDERID
	#Then adds the important data from the XLSX report to the ACE
	consolidated_ACE_data = []
	for json_entry in ACE_data:
		#Match Manifest entries and append relevant data
		for XLSX_line in XLSX_data:
			if XLSX_line[ORDERID_column_index] == json_entry["ORDERID"]: #TODO
				json_entry["client"] = XLSX_line[client_name_column_index]
				json_entry["carrier"] = XLSX_line[carrier_column_index]
				json_entry["closeDate"] = format_date(XLSX_line[ship_date_column_index].split(" ")[0])
				json_entry["trackingNumber"] = XLSX_line[tracking_number_column_index]
		consolidated_ACE_data.append(json_entry)

	return consolidated_ACE_data

def create_trip_summary():
	filename = trip.date + os.sep + trip.date + "-" + trip.BoL + ".info"
	data = {
		"date": trip.date,
		"BoL": trip.BoL,
		"PAPS": trip.PAPS,
		"SCAC": trip.SCAC,
		"gaylord_assignment": trip.gaylord_assignment,
		"gaylord_count": trip.gaylord_count,
		"usps_count": trip.usps_count,
		"dhl_count": trip.dhl_count,
		"fedex_count": trip.fedex_count,
		"package_count": trip.package_count,
		"total_weight": trip.total_weight,
		"shipper": trip.shipper
	}
	with open(filename, "w") as trip_summary_file:
		json.dump(data, trip_summary_file, indent = 4)

	app.setLabel("trip_summary_label", "Trip Summary loaded & ready for emailing")

def load_trip_summary():
	if app.getEntry("Trip.info file:") == "":
		error_box("Please browse and select a .info file")
	else:
		with open(app.getEntry("Trip.info file:")) as trip_summary_file:
			data = json.load(trip_summary_file)
			trip.date = data["date"]
			trip.BoL = data["BoL"]
			trip.PAPS = data["PAPS"]
			trip.SCAC = data["SCAC"]
			trip.gaylord_assignment = data["gaylord_assignment"]
			trip.gaylord_count = data["gaylord_count"]
			trip.usps_count = data["usps_count"]
			trip.dhl_count = data["dhl_count"]
			trip.fedex_count = data["fedex_count"]
			trip.package_count = data["package_count"]
			trip.total_weight = data["total_weight"]
			trip.shipper = data["shipper"]
			
			app.setLabel("trip_summary_label", "Trip Summary loaded & ready for emailing")

	#Put a textbox that updates

def load_ACE():
	"""Loads ACE manifest, returns as a list."""

	ACE_data = []
	try:
		with open(app.getEntry("ACEManifestFileEntry"), "r") as ACE_file:
			ACE_data = json.load(ACE_file)
	except: error_box("Error loading ACE Manifest.\nFile open in another program?")
	return ACE_data

def load_manifest_report():
	"""Loads the manfiest report and returns as a list.

	Also returns ORDERID, Client Name, Carrier, Ship Date, and Tracking Number columns for easy processing
	"""

	XLSX_data = []
	file_path = app.getEntry("XLSXReportFileEntry")
	try:
		if file_path[-5:].lower() == ".xlsx":
			ORDERID_column_index, client_name_column_index, carrier_column_index, ship_date_column_index, tracking_number_column_index = -1, -1, -1, -1, -1
			worksheet = pyxl.load_workbook(file_path).active

			#Load Header
			XLSX_header = ()
			for row in worksheet.values: #Looks odd but openpyxl doesn't like to just _let_ you access the values of the first row
				XLSX_header = row
				break #Effectively sets XLSX_header <= row[0]

			#Figures out where all the important columns are, in order to load that data
			#Can change the names of the important columns in CONFIG.json
			for i, cell in enumerate(XLSX_header):
				if XLSX_header[i] == config.data["XLSX_Report_ORDERID_column_name"]:
					ORDERID_column_index = i
				if XLSX_header[i] == config.data["XLSX_Report_client_name_column_name"]:
					client_name_column_index = i
				if XLSX_header[i] == config.data["XLSX_Report_carrier_column_name"]:
					carrier_column_index = i
				if XLSX_header[i] == config.data["XLSX_Report_ship_date_column_name"]:
					ship_date_column_index = i
				if XLSX_header[i] == config.data["XLSX_Report_tracking_number_column_name"]:
					tracking_number_column_index = i

			#Loads unfiltered Data
			for row in worksheet.values:
				XLSX_data.append(row)

		elif file_path[-4:] == ".csv":
			error_box("Manifest Report in CSV format not yet supported. Please use XLSX file for now")

	except:
		error_box("Error loading XLSX data.\nFile open in another program?")
		traceback.print_exc()

	if XLSX_data == []:
		raise Exception("Empty Manifest Report Data")
		error_box("ERROR!\nNo data read from Manifest Report")
	return XLSX_data, ORDERID_column_index, client_name_column_index, carrier_column_index, ship_date_column_index, tracking_number_column_index

def load_FDA_SKUs():
	"""Loads the FDA_Master_file.XLSX, finds the  column specified by "FDA_Master_file_default_sku_name_column" in the CONFIG.json, and returns as a list."""

	try:
		#NOTE: Currently if a client has FDA-regulated goods, none of their products can go through Section 321 (aka end up on the ACE)
		#This may possibly change in the future
		FDA_SKUs_list = []
		with open("resources/MASTER_FDA_LIST.csv", "r") as SKUs_file:
			csv_reader = csv.reader(SKUs_file, delimiter = ",")
			for line in csv_reader:
				FDA_SKUs_list.append(line[52].strip())
		FDA_SKUs_list.pop(0) #Gets rid of the header
		return FDA_SKUs_list
	except:
		error_box("Error loading resources/MASTER_FDA_LIST.csv.\nFile missing/open in another program?")

def load_batches_file():
	"""Loads the USTM_Scans.XLSX file, and returns the scans as a list.

	Supports CSV and XLSX scans files
	Automatically finds the Scans column

	TODO Check for OS locks
	"""

	if app.getEntry("batchesFileEntry")[-4:].lower() == ".csv":
		try:
			#Loads CSV Data
			csv_data = []
			with open(app.getEntry("batchesFileEntry"), "r") as batches_file:
				csv_reader = csv.reader(batches_file, delimiter = ",")
				for row in csv_reader:
					csv_data.append(row)

			#Reads the ehader to figure out which column is for Batches and which is for the gaylord its assigned to
			csv_header = csv_data[0]
			batch_index = -1
			gaylord_index = -1
			batches_data = []
			for i, cell in enumerate(csv_header):
				if fuzzy_match(cell, config.data["BATCHES_SCANS_batch_column_name"]):
					batch_index = i
				if fuzzy_match(cell, config.data["BATCHES_SCANS_gaylord_column_name"]):
					gaylord_index = i
			#print("Batch Index:", batch_index, "Gaylord Index:" , gaylord_index)
			csv_data.pop(0) #Remove header from CSV data

			#Builds the list of batch-gaylord assignments
			for row in csv_data:
				batches_data.append({"batch": str(row[batch_index]), "gaylord": str(row[gaylord_index])})
			return batches_data
		except:
			error_box("Error loading Scans File.\nFile open in another program?")
			traceback.print_exc()

	elif app.getEntry("batchesFileEntry")[-5:].lower() == ".xlsx": #If user is uploading Detailed Report
		try:
			file_path = app.getEntry("batchesFileEntry")
			workbook = pyxl.load_workbook(file_path)
			worksheet = workbook[config.data["Detailed_Report_scan_sheet_name"]]

			#Load Header from Detailed Report
			header = ()
			for row in worksheet.values:
				header = row
				break

			#Figure out where the batch column is and where the gaylord column is
			batch_index = -1
			gaylord_index = -1
			for i, cell in enumerate(header):
				if cell != None:
					if fuzzy_match(cell, config.data["Detailed_Report_batch_column_name"]):
						batch_index = i
					if fuzzy_match(cell, config.data["Detailed_Report_gaylord_column_name"]):
						gaylord_index = i
				else:
					print(f"Unable to read row {i}")

			#Loads raw XLSX values
			XLSX_data = []
			for row in worksheet.values:
				XLSX_data.append(row)
			XLSX_data.pop(0) # Removes header

			#Builds the list of batch-gaylord assignments
			batches_data = []
			for row in XLSX_data:
				if row[batch_index] != None and row[gaylord_index] != None:
					batches_data.append({"batch": str(row[batch_index]), "gaylord": str(row[gaylord_index])})

			return batches_data

		except:
			error_box("Error loading Scans File.\nFile open in another program?")
			traceback.print_exc()

def construct_ACE(consolidated_json):
	'''Filters the ACE Manifest, returning only entries that match batch/order scans from the Scans File.
	Creates 2 datasets: a Master ACE (used for package counts, detailed report, etc) and a S321_ACE (Uploaded to BorderConnect, contains non-FDA entries)
	'''

	scans_list = load_batches_file()
	FDA_SKUs_list = load_FDA_SKUs()
	master_ACE = [] #Master data, used for package counts/detailed report
	S321_ACE = [] #Master Data, minus the FDA-regulated packages. Sent to BorderConnect
	#good_batches = [] #Used to identify batches that did not match any ACE entries
	good_orders = [] #Used to identify orders that did not match any ACE entries
	warnable_client_found_flag = False

	#Make a batch -> [orders] reference table for use later
	batch_to_order_lookup_dict = {}
	for entry in consolidated_json:
		entry_batchid = entry["BATCHID"]
		if not batch_to_order_lookup_dict.get(entry_batchid):
			batch_to_order_lookup_dict[entry_batchid] = [entry["ORDERID"]]
		else:
			batch_to_order_lookup_dict[entry_batchid].append(entry["ORDERID"])
		#print(f"Adding {entry_batchid} to batch_to_order_lookup_dict")

	#Hashtable lookups to avoid a double for-loop
	searchable_ACE = {entry["ORDERID"]: entry for entry in consolidated_json}

	#Checks if a line is a order or a batch
	for scan in scans_list:
		batch_or_order = scan["batch"]
		orders_list = []
		
		if len(batch_or_order) == 6: #It's a batchid
			#print(f"Looking for orders from batch {batch_or_order}")
			try:
				orders_list = batch_to_order_lookup_dict[batch_or_order]
			except:
				error_box(f"Unable to find {batch_or_order} in batch lookup table (is missing from manifest)")
		elif len(batch_or_order) == 8:
			orders_list = [batch_or_order]
		
		#print(get_orders_from_batch(scan["batch"]))
		#orders_list = get_orders_from_batch(scan["batch"])

		#Matches Batches Scans to ACE entries. Only ACE entries that match a batch scan are added to the out-bound ACE
		master_orders_list = []
		for order in orders_list:
			master_orders_list.append(order)
			if not searchable_ACE.get(order): #If you can find the order
				error_box(f"Unable to find order {order} in ACE Manfiest!\nPlease use Batches page to find the manifest this order belongs to and add that manifest to the ACE")
			else: #Match
				json_entry = searchable_ACE.get(order)
				#Warn if the order's product-seller is on the warnable_clients list
				if json_entry["client"] in [line for line in app.getTextArea("warnableClientsTextArea").split("\n") if line != ""]:
				#if json_entry["client"] in config.data["warnable_clients"]:
					error_box("Entry {} for warnable client {} found".format(json_entry["ORDERID"], json_entry["client"]))
					warnable_client_found_flag = True

				#Makes a temporary entry to validate commodities
				_tmp_entry = json_entry.copy()
				_tmp_entry["commodities"] = []
				all_products_not_FDA_flag = True
				for product in json_entry["commodities"]:
					if product["description"].strip() in FDA_SKUs_list:
						all_products_not_FDA_flag = False
					else:
						_tmp_entry["commodities"].append(product)

				if not all_products_not_FDA_flag:
					_tmp_entry["shipmentClearance"] = "FDA"
					json_entry["shipmentClearance"] = "FDA"
				else:
					_tmp_entry["shipmentClearance"] = "S321"
					json_entry["shipmentClearance"] = "S321"
				_tmp_entry["GAYLORD"] = scan["gaylord"]
				json_entry["GAYLORD"] = scan["gaylord"]
				master_ACE.append(json_entry) #Add to master data (Used for Detailed Report/package counts, etc)
				#good_batches.append(batch_or_order) #Creates a list of good batches, for comparision to original to identify errors
				good_orders.append(order)
				if _tmp_entry["commodities"] != []: #FDA-regulated commodities are removed from the _tmp_entry
					S321_ACE.append(_tmp_entry) #ACE contains only non-FDA-regulated (read: Section 321) entries
					
	if warnable_client_found_flag:
		error_box("Clients that are red-flagged have orders in ACE!\nPlease see Error Log for more details")

	S321_ACE = validate_JSON(S321_ACE, warn_errors_flag=True) #Removes duplicates and common errors that prevent ACE uploading
	master_ACE = validate_JSON(master_ACE, warn_errors_flag=False)

	#Print ACE Manifest
	with open(trip.date + os.sep + trip.date + "-ACE.json", "w") as S321_ACE_file:
		json.dump(S321_ACE, S321_ACE_file, indent = 4)

	report_unmatched_orders(master_orders_list, good_orders)

	return master_ACE, S321_ACE, good_orders

def report_unmatched_orders(orders_list, good_orders):
	'''Compares the list of scans to the list of matched entries, and writes discrepancies to the error-file.
	If a scan is present on the Scans File, said package is on the truck. If it's not in the good_orders, it wasn't matched, and therefore there is no ACE data and the package is undeclared.
	This is bad. Makes an error-box when this occurs, so that end-user can fix
	'''

	#Warn unmatched Batch Scans
	unmatched_orders_list = []
	for order in orders_list:
		if order not in good_orders:
			if order not in unmatched_orders_list:
				unmatched_orders_list.append(order)
	
	#Send out an error_box with unmatched batches
	if unmatched_orders_list != []:
		#print("\n>Printing Unmatched Orders")
		#print("Use the Batches page in Techship to find out what manfiest these orders belong to, and include them in the ACE")
		out_text = "Found Unmatched Orders\nUse the Batches page in Techship to find out what manfiest these belong to, and include them in the ACE (via Manifest History page)"
		#error_file("\n>Printing Unmatched Orders")
		#error_file("\nUse the Orders page in Techship to find out what manfiest these belong to, and include them in the ACE")
		for order in unmatched_orders_list:
			out_line = "Unmatched batch: {}".format(order.ljust(8, " "), order)
			#print(out_line)
			out_text += "\n" + out_line
			#error_file(out_line)
		error_box(out_text)

def validate_JSON(in_json, warn_errors_flag = True):
	"""Removes duplicates from the ACE and cleans up common errors."""

	out_json = []
	orders_list = []
	error_once_flag = False

	# Get rid of duplicates
	for entry in in_json:
		try:
			do_not_ship_flag = False
			commodity_missing_value_flag = False

			#LUS tickets can be split eg. "G1,2"
			#This method assigns the entry to the first gaylord (eg. G1) if it belongs to USPS and the second gaylord (eg. G2) otherwise
			if "," in entry["GAYLORD"]:
				if entry["carrier"] == "EHUB":
					entry["GAYLORD"] = entry["GAYLORD"].split(",")[0]
				else:
					entry["GAYLORD"] = "G" + entry["GAYLORD"].split(",")[1].replace("G", "")
			elif "," in entry["GAYLORD"] and warn_errors_flag:
				if warn_errors_flag: error_box(f"WARNING! Non-LUS pick ticket assigned to multiple gaylords!\n{entry['BATCHID']} {entry['GAYLORD']} {entry['client']}")
			
			#Shipper override
			entry["shipper"] = trip.shipper["shipper"]
			entry["shipmentControlNumber"] = trip.shipper["SCAC"] + entry["shipmentControlNumber"][4:]

			#Clean non-alphanumeric Characters
			entry["consignee"]["name"] = clean_string(entry["consignee"]["name"])
			entry["consignee"]["address"]["addressLine"] = clean_string(entry["consignee"]["address"]["addressLine"])
			entry["consignee"]["address"]["city"] = clean_string(entry["consignee"]["address"]["city"])

			#Validate Name
			if len(entry["consignee"]["name"]) < 2:
				entry["consignee"]["name"] = entry["consignee"]["name"].ljust(3, "A")
			elif len(entry["consignee"]["name"]) > 60:
				entry["consignee"]["name"] = entry["consignee"]["name"][:60]
			#Validate Address
			if len(entry["consignee"]["address"]["addressLine"]) < 2:
				entry["consignee"]["address"]["addressLine"] = entry["consignee"]["address"]["addressLine"].ljust(3, "A")
			elif len(entry["consignee"]["address"]["addressLine"]) > 55:
				entry["consignee"]["address"]["addressLine"] = entry["consignee"]["address"]["addressLine"][:55]
			#Validate City
			if len(entry["consignee"]["address"]["city"]) < 2:
				entry["consignee"]["address"]["city"] = entry["consignee"]["address"]["city"].ljust(3, "A")
			elif len(entry["consignee"]["address"]["city"]) > 30:
				entry["consignee"]["address"]["city"] = entry["consignee"]["address"]["city"][:30]
			#Validate State
			states_list = ["AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY", "AS", "DC", "GU", "MP", "PR", "VI"]
			if entry["consignee"]["address"]["stateProvince"] not in states_list:
				entry["consignee"]["address"]["stateProvince"] = "NY"
			#Validate Zip
			if len(entry["consignee"]["address"]["postalCode"]) != 5:
				entry["consignee"]["address"]["postalCode"] = entry["consignee"]["address"]["postalCode"].ljust(5, "0")

			#Auto-correct a common error (VT is invalid for Vietnam)
			for commodity in entry["commodities"]:
				if commodity.get("countryOfOrigin") == "VT":
					commodity["countryOfOrigin"] = "VN"

			#Find commodities that are missing values
			shipment_value = 0.0
			for commodity in entry["commodities"]:
				for blacklist_entry in config.data["SKUs_blacklist"]:
					if commodity["description"] == blacklist_entry["description"]:
						if trip.shipper["name"] in blacklist_entry["carriers"]:
							for action in blacklist_entry["actions"]:
								if action["type"] == "do_not_ship":
									do_not_ship_flag = True
								elif action["type"] == "error_box" and not error_once_flag:
									error_once_flag = True
									error_box(action["message"])

				if commodity.get("value", False) == False:
					commodity_missing_value_flag = True
					if warn_errors_flag: error_box("Entry {} {} missing commodity value (because it is FDA-regulated or a non-commercial insert).\nPlease add to /resources/FDA_Master_List.csv".format(entry["ORDERID"], entry["client"]))
				else:
					shipment_value += float(commodity.get("value"))

			#Warn if shipment value is over the S321 800$ limit
			if shipment_value >= 800.0:
				do_not_ship_flag = True
				if warn_errors_flag: error_box("Entry {} {} over S321 800$ limit.\nPlease remove from Gaylord!\n(Entry removed from ACE)".format(entry["ORDERID"], entry["client"]))

			#Check for duplicates
			if entry["ORDERID"] not in orders_list and not do_not_ship_flag:
				orders_list.append(entry["ORDERID"])
				out_json.append(entry)
			else:
				if warn_errors_flag: print("Removing duplicate order for {}".format(entry["ORDERID"]))
				# TODO create a "removed duplicates" file
		except:
			traceback.print_exc()
			print("Error at:")
			print(entry)

	return out_json

def assign_gaylords(master_ACE):
	#Reads the Master ACE and 
	unique_gaylords_list = []
	trip.gaylord_assignment = []

	for entry in master_ACE:
		#Clean it first
		if len(entry["GAYLORD"]) == 2:
			entry["GAYLORD"] = "G0" + entry["GAYLORD"][-1]
		#Add it to the uniques list
		match = False
		for unique_gaylord in unique_gaylords_list:
			if entry["GAYLORD"] == unique_gaylord["id"]:
				match = True
		if not match:
			unique_gaylords_list.append({"id": entry["GAYLORD"], "hasFDA": False, "hasUSPS": False, "hasDHL": False, "hasFedex": False, "packages": 0})
	
	#For every gaylord it checks the master file to see where the gaylord is heading and if it has FDA products
	for entry in master_ACE:
		for line in unique_gaylords_list:
			if entry["GAYLORD"] == line["id"]: #Will match only one gaylord
				#Now set flags
				line["packages"] += 1
				if entry["carrier"] in config.data["DHL_carriers_names"]:
					line["hasDHL"] = True
				elif entry["carrier"] in config.data["FEDEX_carriers_names"]:
					line["hasFedex"] = True
				elif entry["carrier"] in config.data["USPS_carriers_names"]:
					line["hasUSPS"]= True
				else:
					error_box("Package with Order ID {} found without carrier".format(entry["ORDERID"]))
				if entry["shipmentClearance"] == "FDA":
					line["hasFDA"] = True

	#Checks if the gaylord has multiple carriers assigned and counts gaylords
	usps_count,	dhl_count, fedex_count = 0, 0, 0
	for line in unique_gaylords_list:
		out = {}
		out["id"] = line["id"]

		carriers_count = 0
		if line["hasUSPS"]: carriers_count += 1
		if line["hasDHL"]: carriers_count += 1
		if line["hasFedex"]: carriers_count += 1

		if carriers_count <= 0:
			error_box("No carriers found for {}?".format(line["id"]))
		if carriers_count > 1:
			error_box("Multiple carriers found for {}".format(line["id"]))
			#TODO error box with option to ignore
		elif line["hasUSPS"]:
			out["carrier"] = "EHUB"
			usps_count += 1
		elif line["hasDHL"]:
			out["carrier"] = "DHLGLOBALMAIL"
			dhl_count += 1
		elif line["hasFedex"]:
			out["carrier"] = "FEDEX"
			fedex_count += 1
			
		if line["hasFDA"]: out["hasFDA"] = "FDA"
		else: out["hasFDA"] = ""

		out["packageCount"] = line["packages"]
		trip.gaylord_assignment.append(out)

	trip.gaylord_assignment.sort(key = lambda x: x["id"])

	trip.gaylord_count = usps_count + dhl_count + fedex_count
	trip.usps_count = usps_count
	trip.dhl_count = dhl_count
	trip.fedex_count = fedex_count
	trip.package_count = len(master_ACE)
	trip.total_weight = int(len(master_ACE) / 2.2)

## Paperwork functions

def create_detailed_report(master_ACE):
	# The report used by Tri-Ad/us to figure out what packages belong to what gaylord/shipping information
	file_name = trip.date + os.sep + trip.date + "-Detailed_Report.csv"
	try:
		with open(file_name, "w", newline = "") as report_file:
			csv_writer = csv.writer(report_file)
			csv_writer.writerow(["Gaylord", "Name", "Address", "City", "State", "Country", "ZIP", "BATCHID", "ORDERID", "SCAC", "Service", "Client", "Close Date", "S321/FDA?", "Total Value", "Commodity 1", "Commodity 2", "Commodity 3", "Commodity 4", "Commodity 5", "Commodity 6", "Commodity 7", "Commodity 8", "Commodity 9", "Commodity 10"])
			for entry in master_ACE:
				#Ugly csv line building. Please ignore
				total_value = 0.0
				commodities_names_list = []
				for commodity in entry["commodities"]:
					commodities_names_list.append(commodity["description"])
					if commodity.get("value"):
						total_value += float(commodity.get("value"))
				out_line = [entry["GAYLORD"], entry["consignee"]["name"], entry["consignee"]["address"]["addressLine"], entry["consignee"]["address"]["city"], entry["consignee"]["address"]["stateProvince"], entry["consignee"]["address"]["country"], entry["consignee"]["address"]["postalCode"], entry["BATCHID"], entry["ORDERID"], entry["shipmentControlNumber"], entry["carrier"], entry["client"], entry["closeDate"], entry["shipmentClearance"], str(total_value)] + commodities_names_list
				csv_writer.writerow(out_line)
	except:
		error_box(f"Unable to create Detailed Report.\nIs file open in another program?")

def create_BoL():
	#Loads a .jpg file template, then writes the last few details to the image before saving it as a .pdf
	file_name = trip.date + os.sep + trip.date + "-Stalco-BoL.pdf"
	image_file_path = "resources/STALCO_BOL.jpg"
	c = canvas.Canvas(file_name, pagesize = (1668, 1986))
	image = ImageReader(image_file_path)
	c.drawImage(image, 0, 0, mask = "auto")

	c.setFont("Courier", 24)
	c.drawString(4, 1628, trip.date)
	c.drawString(1555, 1656, trip.BoL)
	c.drawString(1562, 1631, trip.PAPS)
	c.drawString(134, 776, str(trip.gaylord_count))
	c.drawString(176, 676, str(trip.usps_count))
	c.drawString(171, 626, str(trip.dhl_count))
	c.drawString(181, 576, str(trip.fedex_count))
	c.drawString(525, 811, str(trip.package_count))
	c.drawString(854, 811, str(trip.total_weight))
	#TODO Add shipper profile info (address, etc)
	c.drawString(15, 1521, trip.shipper["shipper"]["name"])
	c.drawString(15, 1425, trip.shipper["shipper"]["address"]["addressLine"])
	c.drawString(900, 1425, (trip.shipper["shipper"]["address"]["city"] + ", " + trip.shipper["shipper"]["address"]["stateProvince"]))
	c.drawString(1160, 1425, trip.shipper["shipper"]["address"]["postalCode"])

	c.showPage()
	try:
		c.save()
	except:
		error_box("Unable to create Stalco BoL.\nIs file open in another program?")

def create_IMS_BoL():
	#Creates the BoL we send to IMS in the same method as the regular BoL
	
	file_name = trip.date + os.sep + trip.date + "-IMS-BoL.pdf"
	image_file_path = "resources/IMS_BOL.jpg"
	c = canvas.Canvas(file_name, pagesize = (1668, 1986))
	image = ImageReader(image_file_path)
	c.drawImage(image, 0, 0, mask = "auto")

	c.setFont("Courier", 36)
	c.drawString(1200, 1700, trip.date)
	c.drawString(1200, 1660, trip.BoL)
	c.drawString(1145, 1190, str(trip.fedex_count + trip.dhl_count))
	c.drawString(240, 544, str(trip.fedex_count) + " pkgs:") #TODO re-add Fedex Package Count
	c.drawString(705, 544, str(trip.dhl_count))

	#Writes a list of FedEx gaylords to the sheet in the proper area
	fedex_gaylords, dhl_gaylords = [], []
	for gaylord in trip.gaylord_assignment:
		if gaylord["carrier"] == "DHLGLOBALMAIL":
			dhl_gaylords.append(gaylord["id"])
		if gaylord["carrier"] == "FEDEX":
			fedex_gaylords.append(gaylord["id"])

	for i, g in enumerate(fedex_gaylords):
		c.drawString(48, 496 - (i * 36), g)
	#Same for DHL
	for i, g in enumerate(dhl_gaylords):
		c.drawString(520, 496 - (i * 36), g)

	c.showPage()
	try:
		c.save()
	except:
		error_box("Unable to create IMS BoL.\nIs file open in another program?")

def create_ProForma(master_ACE):
	#Creates the template we upload to SmartBorder
	#Everything is hard-coded to meet their upload schema. Sorry. 
	try:
		if master_ACE == []: raise Exception
		else: #If loading didn't error out
			commodities_list = {}
			for entry in master_ACE:
				for commodity in entry["commodities"]:
					name = commodity["description"]
					if name not in commodities_list.keys():
						commodities_list[name] = 0 #Add commodity to the list with 0 quantity
					commodities_list[name] = commodities_list[name] + int(commodity["quantity"]) #Update the quantity

			#Techship passes bad data. Error Correction below
			commodities_list = clean_commodities_list(commodities_list)

			fda_list = []
			try:
				with open("resources/MASTER_FDA_LIST.csv", "r") as master_file:
					csv_reader = csv.reader(master_file, delimiter = ",")
					for line in csv_reader:
						if line != "":
							fda_list.append(line)
			except: error_box("Error loading MASTER_FDA_LIST.csv. File open in another program?")

			proforma_lines_data = []
			for commodity in commodities_list:
				for line in fda_list:
					if clean_string(commodity).upper() == clean_string(line[52]).upper() and commodities_list[commodity] != 0 and commodity != "" and line[3] != "NOT SHIPPED": #If commodity description matches description from Master FDA file and there is >0 items
						quantity = commodities_list[commodity]
						price = float(line[3])
						#Ugly CSV line building
						try:
							out = []
							for i in range(0, 50): #line[0] and line[1] are used to match data and don't go on ProForma
								if i == 1:
									out.append(quantity)
								elif i == 3:
									out.append(price)
								elif i == 11:
									out.append(quantity * price)
								elif i == 48:
									out.append(quantity * price)
								else:
									out.append(line[i])
							proforma_lines_data.append(out)
						except:
							traceback.print_exc()
							print("Unable to output to Proforma:", commodity)
			#More ugly
			master_proforma_header = ["ShipperRefNum", "PostToBroker", "InvoiceDate", "StateDest", "PortEntry", "MasterBillofLading", "Carrier", "EstDateTimeArrival", "TermsofSale", "RelatedParties", "ModeTrans", "ExportReason", "FreightToBorder", "ContactInformation", "IncludesDuty", "FreightPrepaidIncluded", "FreightPrepaidNotIncluded", "FreightCollect", "FreightCurrency", "FreightExchange", "IncludesBrokerage", "Currency", "TotalGrossWeightKG", "RailCarNumber", "ShippingQuantity", "ShippingUOM", "DutyandFeesBilledTo", "InvoiceNumber", "OwnerOfGoods", "PurchaseOrder", "EntryIsInBond", "ShipperCustNo", "ShipperName", "ShipperName2Type", "ShipperName2", "ShipperTaxID", "ShipperAddr1", "ShipperAddr2", "ShipperCity", "ShipperState", "ShipperCountry", "ShipperPostalCode", "ShipperMfgID", "ShipToCustNo", "ShipToName", "ShipToName2Type", "ShipToName2", "ShipToTaxID", "ShipToAddr1", "ShipToAddr2", "ShipToCity", "ShipToState", "ShipToCountry", "ShipToPostalCode", "SellerCustNo", "SellerName", "SellerName2Type", "SellerName2", "SellerTaxID", "SellerAddr1", "SellerAddr2", "SellerCity", "SellerState", "SellerCountry", "SellerPostalCode", "MfgCustNo", "MfgName", "MfgName2Type", "MfgName2", "MfgID", "MfgAddress1", "MfgAddress2", "MfgCity", "MfgState", "MfgCountry", "MfgPostalCode", "BuyerCustNo", "BuyerName", "BuyerName2Type", "BuyerName2", "BuyerUSTaxID", "BuyerAddress1", "BuyerAddress2", "BuyerCity", "BuyerState", "BuyerCountry", "BuyerPostalCode", "ConsigneeCustNo", "ConsigneeName", "ConsigneeName2Type", "ConsigneeName2", "ConsigneeUSTaxID", "ConsigneeAddress1", "ConsigneeAddress2", "ConsigneeCity", "ConsigneeState", "ConsigneeCountry", "ConsigneePostalCode", "PartNumber", "Quantity", "QuantityUOM", "UnitPrice", "GrossWeightKG", "NumberOfPackages", "PackageUOM", "CountryOrigin", "SPI", "ProductClaimCode", "Description", "ValueOfGoods", "LineMfgCustNo", "LineMfgName", "LineMfgName2Type", "LineMfgName2", "LineMfgID", "LineMfgAddress1", "LineMfgAddress2", "LineMfgCity", "LineMfgState", "LineMfgCountry", "LineMfgPostalCode", "LineBuyerCustNo", "LineBuyerName", "LineBuyerName2Type", "LineBuyerName2", "LineBuyerUSTaxID", "LineBuyerAddress1", "LineBuyerAddress2", "LineBuyerCity", "LineBuyerState", "LineBuyerCountry", "LineBuyerPostalCode", "LineConsigneeSameAsBuyer", "LineConsigneeCustNo", "LineConsigneeName", "LineConsigneeName2Type", "LineConsigneeName2", "LineConsigneeUSTaxID", "LineConsigneeAddress1", "LineConsigneeAddress2", "LineConsigneeCity", "LineConsigneeState", "LineConsigneeCountry", "LineConsigneePostalCode", "LineNote", "Tariff1Number", "Tariff1ProductValue", "Tariff1Quantity1", "Tariff1Quantity1UOM", "Tariff1Quantity2", "Tariff1Quantity2UOM", "Tariff1Quantity3", "Tariff1Quantity3UOM", "Tariff2Number", "Tariff2ProductValue", "Tariff2Quantity1", "Tariff2Quantity1UOM", "Tariff2Quantity2", "Tariff2Quantity2UOM", "Tariff2Quantity3", "Tariff2Quantity3UOM", "Tariff3Number", "Tariff3ProductValue", "Tariff3Quantity1", "Tariff3Quantity1UOM", "Tariff3Quantity2", "Tariff3Quantity2UOM", "Tariff3Quantity3", "Tariff3Quantity3UOM", "Tariff4Number", "Tariff4ProductValue", "Tariff4Quantity1", "Tariff4Quantity1UOM", "Tariff4Quantity2", "Tariff4Quantity2UOM", "Tariff4Quantity3", "Tariff4Quantity3UOM", "Tariff5Number", "Tariff5ProductValue", "Tariff5Quantity1", "Tariff5Quantity1UOM", "Tariff5Quantity2", "Tariff5Quantity2UOM", "Tariff5Quantity3", "Tariff5Quantity3UOM", "Tariff6Number", "Tariff6ProductValue", "Tariff6Quantity1", "Tariff6Quantity1UOM", "Tariff6Quantity2", "Tariff6Quantity2UOM", "Tariff6Quantity3", "Tariff6Quantity3UOM"]
			master_line_start = [
				trip.BoL, #A
				"FALSE", 
				trip.date,
				"NY", 
				"0901", 
				trip.PAPS, 
				trip.SCAC, 
				trip.date.replace("-", "/") + " 03:00 PM",
				"PLANT",
				"",
				"30", #K
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"USD", #X
				int(trip.total_weight),
				"", #Z
				trip.package_count,
				"PCS",
				"Buyer",
				"",
				"",
				"",
				"",
				"",
				"STALCO INC", #AI
				"",
				"",
				"160901-55044",
				"401 CLAYSON RD",
				"",
				"NORTH YORK",
				"ON",
				"CA",
				"M9M 2H4",
				"XOSTAINC401NOR",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"",
				"STALCO INC", #BF
				"",
				"",
				"160901-55044",
				"401 CLAYSON RD",
				"",
				"NORTH YORK",
				"ON",
				"CA",
				"M9M 2H4",
				"",
				"STALCO INC", #BQ
				"",
				"",
				"XOSTAINC401NOR",
				"401 CLAYSON RD",
				"",
				"NORTH YORK",
				"ON",
				"CA",
				"M9M 2H4",
				"",
				"IMS OF WESTERN NY", #CB
				"",
				"",
				"16-131314301",
				"2540 WALDEN AVE",
				"SUITE 450",
				"BUFFALO",
				"NY",
				"US",
				"14225",
				"",
				"IMS OF WESTERN NY", #CM
				"",
				"",
				"16-131314301",
				"2540 WALDEN AVE",
				"SUITE 450",
				"BUFFALO",
				"NY",
				"US",
				"14225" #CV
			]
			
			proforma_lines_header = ["PartNumber", "Quantity", "QuantityUOM", "UnitPrice", "GrossWeightKG", "NumberOfPackages", "PackageUOM", "CountryOrigin", "SPI", "ProductClaimCode", "Description", "ValueOfGoods", "LineMfgCustNo", "LineMfgName", "LineMfgName2Type", "LineMfgName2", "LineMfgID", "LineMfgAddress1", "LineMfgAddress2", "LineMfgCity", "LineMfgState", "LineMfgCountry", "LineMfgPostalCode", "LineBuyerCustNo", "LineBuyerName", "LineBuyerName2Type", "LineBuyerName2", "LineBuyerUSTaxID", "LineBuyerAddress1", "LineBuyerAddress2", "LineBuyerCity", "LineBuyerState", "LineBuyerCountry", "LineBuyerPostalCode", "LineConsigneeSameAsBuyer", "LineConsigneeCustNo", "LineConsigneeName", "LineConsigneeName2Type", "LineConsigneeName2", "LineConsigneeUSTaxID", "LineConsigneeAddress1", "LineConsigneeAddress2", "LineConsigneeCity", "LineConsigneeState", "LineConsigneeCountry", "LineConsigneePostalCode", "LineNote", "Tariff1Number", "Tariff1ProductValue", "Tariff1Quantity1", "Tariff1Quantity1UOM", "Tariff1Quantity2", "Tariff1Quantity2UOM", "Tariff1Quantity3", "Tariff1Quantity3UOM", "Tariff2Number", "Tariff2ProductValue", "Tariff2Quantity1", "Tariff2Quantity1UOM", "Tariff2Quantity2", "Tariff2Quantity2UOM", "Tariff2Quantity3", "Tariff2Quantity3UOM", "Tariff3Number", "Tariff3ProductValue", "Tariff3Quantity1", "Tariff3Quantity1UOM", "Tariff3Quantity2", "Tariff3Quantity2UOM", "Tariff3Quantity3", "Tariff3Quantity3UOM", "Tariff4Number", "Tariff4ProductValue", "Tariff4Quantity1", "Tariff4Quantity1UOM", "Tariff4Quantity2", "Tariff4Quantity2UOM", "Tariff4Quantity3", "Tariff4Quantity3UOM", "Tariff5Number", "Tariff5ProductValue", "Tariff5Quantity1", "Tariff5Quantity1UOM", "Tariff5Quantity2", "Tariff5Quantity2UOM", "Tariff5Quantity3", "Tariff5Quantity3UOM", "Tariff6Number", "Tariff6ProductValue", "Tariff6Quantity1", "Tariff6Quantity1UOM", "Tariff6Quantity2", "Tariff6Quantity2UOM", "Tariff6Quantity3", "Tariff6Quantity3UOM"]

			#Outputs the excel file for SmartBorder upload
			workbook = pyxl.Workbook()
			filename = trip.date + os.sep + trip.date + "-ProForma_Template.xlsx"
			worksheet = workbook.active
			worksheet.title = "Sheet1"
			worksheet.append(master_proforma_header)
			for row in proforma_lines_data:
				worksheet.append(master_line_start + row)
			workbook.save(filename)

			#USGR data, for easy USGR generation
			workbook = pyxl.Workbook()
			filename = trip.date + os.sep + trip.date + "-USGR_Data.xlsx"
			worksheet = workbook.active
			worksheet.title = "Sheet1"
			worksheet.append(proforma_lines_header)
			for row in proforma_lines_data:
				worksheet.append(row)
			workbook.save(filename)

			'''
			#Creates the USGR Data file which is used by USTM to make USGRs
			with open(config.data["default_date"] + os.sep + config.data["default_date"] + "-USGR_Data.csv", "w", newline = "") as USGR_file:
				csv_writer = csv.writer(USGR_file)
				for row in proforma_lines_data:
					csv_writer.writerow(row)
			'''
	except:
		error_box("Unable to create ProForma.\nIs file open in another program?")

def create_loadsheet():
	#Creates a useful but non-essential sheet detailing each gaylord, where its going, if it has FDA products, and how many packages it holds
	global master_metadata
	try:
		#Write to file
		file_name = trip.date + os.sep + trip.date + "-Load_Sheet.pdf"
		c = canvas.Canvas(file_name, pagesize = (595.27, 841.89), bottomup = 0)
		c.setFont("Courier", 11)

		c.drawString(10, 22, "ROW | LEFT | RIGHT|           " + app.getEntry("Date:") + " Load Sheet")
		for i in range(1, 15):
			c.drawString(10, 8 + (i * 28), " {} |      |      |".format(str(i).ljust(2, " ")))
			c.drawString(10, 22 + (i * 28), "____|______|______|")

		c.drawString(10, 442, "SKID  CARRIER       FDA? PACKAGES")
		for i, row in enumerate(trip.gaylord_assignment):
			c.drawString(10, (456 + i * 14), (row["id"].ljust(6, " ") + row["carrier"].ljust(14, " ") + row["hasFDA"].ljust(5, " ") + str(row["packageCount"])))
		c.showPage()
		c.save()
	except Exception:
		traceback.print_exc()
		error_box("Unable to create Load Sheet.\nIs file open in another program?")

def create_detailed_report_CSV(master_ACE):
	header = ["SCN", "Shipment Type", "Province Of Loading", "Shipper Name", "Shipper Address", "Shipper City", "Shipper State", "Shipper Zip", "Consignee Name", "Consignee Address", "Consignee City", "Consignee State", "Consignee Zip", "Commodity Description", "Commodity Value", "Commodity Quantity", "Commodity Quantity Unit", "Commodity Weight", "Commodity Weight Unit", "Commodity Country of Origin"]
	with open(trip.date + os.sep + trip.date + "-Detailed_Report_CSV.csv", "w", newline = "") as file:
		csv_writer = csv.writer(file)
		csv_writer.writerow(header)
		for entry in master_ACE:
			for i, commodity in enumerate(entry["commodities"]):
				try:
					row = [
						entry["shipmentControlNumber"],
						#increment_SCAC(entry["shipmentControlNumber"], i),
						entry["type"],
						entry["provinceOfLoading"],
						entry["shipper"]["name"],
						entry["shipper"]["address"]["addressLine"],
						entry["shipper"]["address"]["city"],
						entry["shipper"]["address"]["stateProvince"],
						(entry["shipper"]["address"]["postalCode"]).ljust(5, "0"),
						entry["consignee"]["name"],
						entry["consignee"]["address"]["addressLine"],
						entry["consignee"]["address"]["city"],
						entry["consignee"]["address"]["stateProvince"],
						(entry["consignee"]["address"]["postalCode"]).ljust(5, "0"),
						commodity["description"],
						commodity["value"],
						int(float(commodity["quantity"])),
						commodity["packagingUnit"],
						commodity["weight"],
						commodity["weightUnit"],
						commodity["countryOfOrigin"]
					]
					csv_writer.writerow(row)
				except Exception:
					traceback.print_exc()
					order_id = entry["ORDERID"]
					print(f"Unable to add {order_id} to AC ACE")
					print(json.dumps(entry, indent = 4))

def email_paperwork(button):
	#Check if all files exist
	#TODO some sort of loading for Trip?
	if trip.BoL == None:
		error_box("No trip info found! Make sure to create the paperwork for today, or load a previosuly generated trip for emailing!")
	else:
		folder = trip.date + os.sep + trip.date
		if os.path.exists(folder + "-ACE.json") and \
		   os.path.exists(folder + "-Detailed_Report.csv") and \
		   os.path.exists(folder + "-IMS-BoL.pdf") and \
		   os.path.exists(folder + "-Load_Sheet.pdf") and \
		   os.path.exists(folder + "-Stalco-BoL.pdf") and \
		   os.path.exists(folder + "-Detailed_Report_CSV.csv"):
			try:
				#Get files
				files = []
				files.append(folder + "-ACE.json")
				files.append(folder + "-Detailed_Report.csv")
				files.append(folder + "-IMS-BoL.pdf")
				files.append(folder + "-Load_Sheet.pdf")
				files.append(folder + "-Stalco-BoL.pdf")
				if button == "Email to AmeriConnect":
					files.append(folder + "-AmeriConnect_ACE.csv")
				if app.getEntry("ProFormaFileEntry") != "" and app.getEntry("ProFormaFileEntry")[-4:].lower() == ".pdf":
					files.append(app.getEntry("ProFormaFileEntry"))
				
				#Make email
				username = app.stringBox("Username?", "Please enter your Outlook email address:", parent = None) #Not saved on purpose for security reasons
				password = app.stringBox("Password?", "Please enter the password for {}:".format(username), parent = None)
				
				smtp = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)
				smtp.starttls()
				smtp.login(username, password)

				message = MIMEMultipart()
				message_text = str(app.getTextArea("EmailTextArea")) + "\n" + str(trip.gaylord_count) + " Gaylords total:\n" + str(trip.usps_count) + " for USPS, " + str(trip.dhl_count) + " for DHL, " + str(trip.fedex_count) + " for FedEx\n" + "Please hit \"REPLY ALL\" when responding to this email trail."
				
				message["From"] = username
				recipients_list = trip.shipper["email_list"]
				message["To"] = ", ".join(recipients_list)
				message["Subject"] = trip.date + " " + app.getEntry("Email Subject:")
				message.attach(MIMEText(message_text, "plain"))

				#Attach files to email
				for path in files:
					part = MIMEBase('application', "octet-stream")
					with open(path, 'rb') as file:
						part.set_payload(file.read())
					encoders.encode_base64(part)
					part.add_header('Content-Disposition', 'attachment; filename="{}"'.format(Path(path).name))
					message.attach(part)

				smtp.send_message(message)
				del message
				smtp.quit()
				info_box("Email successfully sent!")
			except:
				error_box("Email failed!")
				traceback.print_exc()
		else:
			error_box("Not all files required for email are present! Did you forget the ProForma or delete a file?")

## Page 2 Functions

def load_ACE_manifest():
	#Loads the ACE so the rest of the features can use the data
	#Also says how many packages are on the ACE as a "green light"
	try:
		with open(app.getEntry("ACEManifestFileEntry2"), "r") as ACE_file:
			ACE_data = json.load(ACE_file)
			app.setLabel("ACE_status_label", "{} ACE Entries Loaded".format(str(len(ACE_data))))
			SCN_ending = ACE_data[0]["shipmentControlNumber"][-2:]
			app.setLabel("SCNLabel", "SCNs currently end with: {}".format(SCN_ending))
			return ACE_data
	except FileNotFoundError: error_box("No file entered")
	except: error_box("Uploaded file was not an ACE Manfiest")

def remove_gaylord():
	#Removes a specific gaylord from the ACE, then prints
	try:
		ACE_data = load_ACE_manifest()
		gaylord = app.getEntry('Gaylord (eg. "G1"):').upper()
		good_entries = []
		bad_entries = []
		try:
			for entry in ACE_data:
				if entry["GAYLORD"] != gaylord:
					good_entries.append(entry)
				else:
					bad_entries.append(entry)
		except:
			error_box("No enties assigned to Gaylord {}".format(gaylord))

		if len(bad_entries) > 0:
			ACE_data = good_entries
			with open(app.getEntry("ACEManifestFileEntry2"), "w") as ACE_file:
				json.dump(ACE_data, ACE_file, indent = 4)
			with open(app.getEntry("ACEManifestFileEntry2") + "-REMOVED_GAYLORDS", "w") as ACE_file:
				json.dump(bad_entries, ACE_file, indent = 4)
			info_box("Gaylord {} successfully removed".format(gaylord))
			app.setLabel("ACE_status_label", "{} ACE Entries Loaded".format(str(len(ACE_data))))
		else:
			error_box("No entries found in Gaylord {}".format(gaylord))
	except:
		error_box("Error removing Gaylord {}. Make sure it matches the name on the Load Sheet (Eg. 'G02')".format(gaylord))

def remove_items():
	#Removes any packages from the ACE that match the BATCHID/ORDERID/SCAC Code as requested
	try:
		ACE_data = load_ACE_manifest()
		batches = app.getTextArea('batchesTextArea').replace("\n", ",").split(",")
		good_entries = []
		bad_entries = []
		for entry in ACE_data:
			matched_flag = False
			for line in batches:
				if entry["BATCHID"] == line or entry["ORDERID"] == line or entry["shipmentControlNumber"] == line: #If Entry matches one of the order_ids to remove
					matched_flag = True
					if entry not in bad_entries:
						bad_entries.append(entry)
			if not matched_flag:
				if entry not in good_entries:
					good_entries.append(entry)

		ACE_data = good_entries
		filename = str(app.getEntry("ACEManifestFileEntry2"))
		with open(filename, "w") as ACE_file:
			json.dump(ACE_data, ACE_file, indent = 4)
		with open(filename + "-REMOVED_ORDERS", "w") as ACE_file:
			json.dump(bad_entries, ACE_file, indent = 4)
		info_box(f"{str(len(bad_entries))} entries removed.\n{str(len(good_entries))} good entries outputted to {filename}")
		app.setLabel("ACE_status_label", "{} ACE Entries Loaded".format(str(len(ACE_data))))

	except:
		error_box("Error removing specified batches/orders. Please check batches and if ACE was loaded")

def change_SCNs():
	try:
		ACE_data = load_ACE_manifest()
		new_SCN = app.getEntry("New 2 digits:")[:2]
		for entry in ACE_data:
			entry["shipmentControlNumber"] = entry["shipmentControlNumber"][:14] + new_SCN
		with open(app.getEntry("ACEManifestFileEntry2"), "w") as ACE_file:
			json.dump(ACE_data, ACE_file, indent = 4)
			app.setLabel("SCNLabel", "SCNs currently end with: {}".format(new_SCN))
		info_box("SCN ending changed to {}".format(new_SCN))
	except:
		error_box("Error changing SCN. Please check new SCN and if ACE was loaded")

def split_ACE():
	try:
		ACE_data = load_ACE_manifest()
		max_ACE_entries = 9999
		for i in range(len(ACE_data) // max_ACE_entries + 1):
			with open("Split_ACE_Manifest_" + str(i) + ".json", "w") as out_file:
				start = i * max_ACE_entries
				end = (i + 1) * max_ACE_entries - 1
				#print(start, end)
				json.dump(ACE_data[start:end], out_file)
		info_box("ACE successfully split!\nPlease check root folder")
	except:
		error_box("Unable to split ACE Manifest. Did you remember to load the manifest?")

## Page 3 Functions

def convert_JSON_to_CSV():
	#Process JSON file
	try:
		filename = app.getEntry("JSON")
		with open(filename) as json_file:
			data = json.load(json_file)
			filepath = filename.replace(".json", ".csv")
		data_file = open(filepath, "w", newline="")
		csv_writer = csv.writer(data_file)

		for l in data: #for line in data:
			try:
				_consignee_province = l["consignee"]["address"]["stateProvince"]
				_consignee_postal_code = l["consignee"]["address"]["postalCode"]
			except:
				_consignee_province = ""
			try:
				_shipper_name = l["shipper"]["name"]
				_shipper_address = l["shipper"]["address"]["addressLine"]
				_shipper_country = l["shipper"]["address"]["country"]
				_shipper_city = l["shipper"]["address"]["city"]
				_shipper_province = l["shipper"]["address"]["stateProvince"]
				_shipper_postal_code = l["shipper"]["address"]["postalCode"]
			except:
				_shipper_name = "Stalco Inc."
				_shipper_address = "401 Clayson Road"
				_shipper_country =  "CA"
				_shipper_city = "Toronto"
				_shipper_province = "ON"
				_shipper_postal_code = "M9M2H4"
			try:
				_client = l["client"]
				_carrier = l["carrier"]
				_closeDate = l["closeDate"]
				_trackingNumber = l["trackingNumber"]
				_gaylord = l["GAYLORD"]
			except:
				_client = "N/A"
				_carrier = "N/A"
				_closeDate = "N/A"
				_trackingNumber = "N/A"
				_gaylord = "N/A"
			
			head = ( #Doing it manually for now. This format doesn't change often
				l["ORDERID"],
				l["BATCHID"],
				l["data"],
				l["type"],
				l["shipmentControlNumber"],
				# Defaults for when ACE is missing entries
				#_province_of_loading,
				"ON", #2020-11-26 hardcoding Temporarily
				_shipper_name,
				_shipper_address,
				_shipper_country,
				_shipper_city,
				_shipper_province,
				_shipper_postal_code,
				l["consignee"]["name"],
				l["consignee"]["address"]["addressLine"],
				l["consignee"]["address"]["country"],
				l["consignee"]["address"]["city"],
				_consignee_province,
				_consignee_postal_code,
				_client,
				_carrier,
				_closeDate,
				_trackingNumber,
				_gaylord)
			for i, commodity in enumerate(l["commodities"]): #for commodity in line["commodities"]
				body = (
					l["commodities"][i]["description"],
					l["commodities"][i]["quantity"],
					l["commodities"][i]["packagingUnit"],
					l["commodities"][i]["weight"],
					l["commodities"][i]["weightUnit"])
				if "value" in l["commodities"][i].keys():
					body = body + (l["commodities"][i]["value"],)
				if "countryOfOrigin" in l["commodities"][i]:
					body = body + (l["commodities"][i]["countryOfOrigin"],)
				row = head + body
				csv_writer.writerow(row)

		data_file.close()
		app.infoBox("Done", f"Finished converting JSON.\nOutputting to {filepath}")
	except:
		error_box("Error converting JSON to CSV.\nPlease see console for more details")

def convert_CSV_to_JSON():
	try:
		filename = app.getEntry("CSV")
		with open(filename) as csv_file:
			csv_reader = csv.reader(csv_file, delimiter = ',')
			csv_data = []
			for row in csv_reader:
				csv_data.append(row)
				
			#Makes a list of consignees to add to JSON
			consignees = []
			for row in csv_data:
				if row[4] not in consignees:
					consignees.append(row[4])

			#For each consignee, add each entry to JSON
			out_json = []
			for consignee in consignees:
				entry = {}
				for row in csv_data:
					if consignee == row[4]:
						entry = {
							"ORDERID": row[0], #I don't care if this is hard-coded. I'm the one who made the format of the CSV being read
							"BATCHID": row[1],
							"data": row[2],
							"type": row[3],
							"shipmentControlNumber": row[4],
							"provinceOfLoading": row[5],
							"shipper": {
								"name": row[6],
								"address": {
									"addressLine": row[7],
									"country": row[8],
									"city": row[9],
									"stateProvince": row[10],
									"postalCode": row[11].zfill(5)
								}
							},
							"consignee": {
								"name": row[12],
								"address": {
									"addressLine": row[13],
									"country": row[14],
									"city": row[15],
									"stateProvince": row[16],
									"postalCode": row[17].zfill(5)
								}
							},
							"client": row[18],
							"carrier": row[19],
							"closeDate": row[20],
							"trackingNumber": row[21],
							"GAYLORD": row[22],
							"commodities": []
						}

				for row in csv_data: #Searches for commodities that match consignee
					if consignee == row[4]:
						commodity = {}
						if len(row) == 30: #If the entry has value and countryOfOrigin
							commodity = {
								"description": row[23],
								"quantity": float(row[24]),
								"packagingUnit": row[25],
								"weight": int(row[26]),
								"weightUnit": row[27],
								"value": row[28],
								"countryOfOrigin": row[29]
							}
						elif len(row) == 29: #If it has only value
							commodity = {
								"description": row[23],
								"quantity": float(row[24]),
								"packagingUnit": row[25],
								"weight": int(row[26]),
								"weightUnit": row[27],
								"value": row[28],
							}
						else:
							commodity = {
								"description": row[23],
								"quantity": float(row[24]),
								"packagingUnit": row[25],
								"weight": int(row[26]),
								"weightUnit": row[27]
							}
						entry["commodities"].append(commodity)
				out_json.append(entry)
				filepath = filename.replace(".csv", ".json")
			with open(filepath, "w") as json_file:
				json.dump(out_json, json_file, indent=4)
			app.infoBox("Done", "Done converting CSV to JSON.\nOutputting to {}".format(filepath))
	except:
		error_box("Error converting from CSV to JSON.\nPlease see console for more details")

def json_beautifier():
	#Adds indents to the JSON to make it human-readable
	try:
		json_file_name = app.getEntry("Ugly JSON")
		with open(json_file_name, "r") as json_file:
			json_data = json.load(json_file)
		with open(json_file_name, "w") as json_file:
			json.dump(json_data, json_file, indent = 4)
		app.infoBox("Done", "Done formatting JSON.\nOutputting to {}".format(json_file_name))
	except:
		error_box("Error beautifying JSON. Make sure uploaded file was valid JSON")
	
def combine_JSONs():
	#Combines 2 JSONs
	out_data = []
	with open(app.getEntry("JSON 1"), "r") as json_file_1:
		json_data_1 = json.load(json_file_1)
	with open(app.getEntry("JSON 2"), "r") as json_file_2:
		json_data_2 = json.load(json_file_2)
	for line in json_data_1:
		out_data.append(line)
	for line in json_data_2:
		out_data.append(line)
	filename = config.data["default_date"] + os.sep + "(Combined)_ACE_Manifest.json"
	with open(filename, "w") as json_file:
		json.dump(out_data, json_file, indent = 4)
	app.infoBox("Done", "Done combining JSONs.\nOutputting to " + filename)

## USGR Stuff

def create_USGR(calling_button = None, USGR_date = None, USGR_BoL = None, USGR_entry = None, USGR_data_file = None):
	"""USGR is paperwork for all the FDA-regulated products that return to the US. For Customs purposes or something."""

	if USGR_date == None:
		USGR_date = "".join(character for character in app.getEntry("USGR Date:") if character in "1234567890-/")
	if USGR_BoL == None:
		USGR_BoL = "".join(character for character in app.getEntry("USGR BoL #:") if character in "1234567890")
	if USGR_entry == None:
		USGR_entry = "".join(character for character in app.getEntry("USGR Entry Number:") if character in "L1234567890-")

	USGR_data = []
	if USGR_data_file == None:
		#Loads the SKUs + Quantities from the USGR Data file or ProForma Lines
		if app.getEntry("USGR Data:")[-4:].lower() == ".csv":
			with open(app.getEntry("USGR Data:"), "r") as proforma_file:
				csv_reader = csv.reader(proforma_file, delimiter = ',')
				for row in csv_reader:
					USGR_data.append(row)
		elif app.getEntry("USGR Data:")[-5:].lower() == ".xlsx":
			USGR_data = load_XLSX_file(app.getEntry("USGR Data:"))
		else:
			error_box("ERROR!\nUnable to open file!")
	else:
		if USGR_data_file[-4:].lower() == ".csv":
			with open("USGR_files/" + USGR_data_file, "r") as proforma_file:
				csv_reader = csv.reader(proforma_file, delimiter = ',')
				for row in csv_reader:
					USGR_data.append(row)
		elif USGR_data_file[-5:].lower() == ".xlsx":
			USGR_data = load_XLSX_file("USGR_files/" + USGR_data_file)

	#Loads the FDA File
	FDA_data = []
	with open("resources/MASTER_FDA_LIST.csv", "r") as fda_file:
		csv_reader = csv.reader(fda_file, delimiter = ',')
		for row in csv_reader:
			FDA_data.append(row)

	#Matches the USGR Data to the Master_FDA_List
	out_data = []
	for commodity in USGR_data:
		for row in FDA_data:
			if commodity[0] == row[0]:
				#print(row)
				out_line = row
				out_line[1] = commodity[1]
				out_line[11] = commodity[11]
				out_data.append(out_line)
	
	#Creates the main USGR information chart
	file_name = "USGR" + os.sep + USGR_date + "-USGR_Table-" + USGR_entry + ".pdf"
	c = canvas.Canvas(file_name, pagesize = (595.27, 841.89), bottomup = 1)
	c.setFont("Courier", 7)
	c.drawString(10, 820, "US GOODS RETURNED")
	c.drawString(10, 812, "INVOICE REFERENCE #: {}    ENTRY #: {}    DATE: {}".format(USGR_BoL, USGR_entry, USGR_date))
	c.drawString(10, 800, "PART/ITEM #       DESCRIPTION                     MANUFACTURER           CITY, STATE           VALUE     QTY  IMPORTDATE  ENTRY PORT")
	write_row_index = 1
	for row in out_data:
		if row[7] == "US":
			out = ""
			out += str(row[0])[:17].ljust(18, " ")
			out += str(row[10])[:30].ljust(32, " ")
			out += str(row[13])[:22].ljust(23, " ")
			out += (str(row[19])[:17] + ", " + str(row[20][:2])).ljust(22, " ")
			out += str(row[11])[:9].ljust(10, " ")
			out += str(row[1])[:4].ljust(5, " ")
			out += get_last_import_date(row[0], USGR_date).ljust(12, " ")
			out += "BUFFALO US"
			c.drawString(10, (800 - (write_row_index * 8)), str(out))
			write_row_index += 1
	c.showPage()

	#Attaches all 6 of the the USGR documents. Adds Entry Number + BoL Number to some of the pages
	c.setFont("Courier", 12)
	for i in range(2, 10): #2 through 9 inclusive
		filepath = "resources" + os.sep + "page_" + str(i) + ".jpg"
		image = Image.open(filepath)
		c.drawImage(ImageReader(image), 0, 0, 595.27, 841.89)
		if i == 2:
			c.drawString(180, 635, USGR_date)
			c.drawString(180, 615, USGR_entry)
			c.drawString(180, 595, USGR_BoL)
		if i == 7:
			c.drawString(140, 612, USGR_entry)
			c.drawString(140, 600, USGR_BoL)
			c.drawString(150, 295, USGR_date)
		c.showPage()

	c.save()
	print(f"USGR for {USGR_date} completed")

def load_XLSX_file(file_path = None):
	try:
		if file_path == None:
			file_path = app.getEntry("USGR Data:")
		workbook = pyxl.load_workbook(file_path)
		worksheet = workbook.active

		#Loads raw XLSX values
		XLSX_data = []
		for row in worksheet.values:
			XLSX_data.append(row)

		return XLSX_data

	except:
		error_box("Error loading Detailed Report.\nFile open in another program?")

def get_token():
	if not config.data.get("token"):
		access_token = generate_token()
		return access_token
	else:
		#Checks if token is expired
		creation_time = datetime.strptime(config.data["token"]["creation_time"], "%Y-%m-%d %H:%M:%S")
		token_duration = config.data["token"]["contents"]["expires_in"]
		time_delta = datetime_module.timedelta(seconds = token_duration)
		target_time = creation_time + time_delta
		now = datetime.now()
		#print(f"TARGET: {target_time} NOW: {now}")

		if now > target_time:
			access_token = generate_token()
			return access_token

def generate_token():
	host_url = "https://secure-wms.com/AuthServer/api/Token"
	headers = {
		"Content-Type": "application/json; charset=utf-8",
		"Accept": "application/json",
		"Host": "secure-wms.com",
		"Accept-Language": "Content-Length",
		"Accept-Encoding": "gzip,deflate,sdch",
		"Authorization": "Basic " + config.data["auth_key"]
	}
	payload = json.dumps({
		"grant_type": "client_credentials",
		"tpl": config.data["tpl"],
		"user_login_id": config.data["user_login_id"]
	})

	responce = requests.request("POST", host_url, data = payload, headers = headers, timeout = 3.0)

	#print("STATUS:", responce.status_code)
	#print(responce.text)
	access_token = responce.json()

	config.data["token"]["contents"] = access_token
	config.data["token"]["creation_time"] = datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")
	config.save()
	print("3PLC Access token renewed")

	return access_token

def get_sku(sku_name, date_in):
	pgsiz = 1000
	pgnum = 1
	total_pages = 1
	out_list = []

	try:
		url = f"https://secure-wms.com/inventory?pgsiz={pgsiz}&pgnum={pgnum}&rql=itemIdentifier.sku=={sku_name};receivedDate=lt={date_in}"
		headers = {
			"Host": "secure-wms.com",
			"Content-Type": "application/hal+json; charset=utf-8",
			"Accept": "application/hal+json",
			"Authorization": "Bearer " + config.data["token"]["contents"]["access_token"],
			"Accept-Encoding": "gzip,deflate,sdch",
			"Accept-Language": "en-US,en;q=0.8",
		}
		payload = {}
		responce = requests.request("GET", url, headers=headers, data=payload)
		responce_json = responce.json()
		#print(json.dumps(responce_json, indent = 4))
		#
		total_results = responce_json["totalResults"]
		total_pages = math.ceil(total_results / pgsiz)
		pgnum += 1
		if responce_json.get("_embedded"):
			if responce_json.get("_embedded").get("item"):
				for item in responce_json["_embedded"]["item"]:
					out_list.append(item)
		print("{} receivers for {} retrieved from 3PLC".format(total_results, sku_name))
	except:
		print("ERROR:")
		print(responce.text)
	return out_list

def find_oldest_receiver_date(orders_list):
	sort_list = []
	for order in orders_list:
		#location = order["locationIdentifier"]["nameKey"]["name"]
		#if "TEST" not in location and "TRIAGE" not in location:
		sort_list.append(order["receivedDate"])
	if len(sort_list) != 0:
		return max(sort_list)[:10] #TMP
	else:
		return "N/A"

def get_last_import_date(sku_name, date_in):
	get_token()
	results_list = get_sku(sku_name, date_in)
	return find_oldest_receiver_date(results_list)

## Utility Functions

def increment_SCAC(SCAC_code: str, i: int):
	carrier_code = SCAC_code[:4]
	numeric_part = SCAC_code[4:]
	numeric_part = str(int(numeric_part) + i)
	return (carrier_code + numeric_part)

def fuzzy_match(value, target):
	if type(target) == list and type(target) != str:
		result = False
		for element in target:
			if fuzzy_match(value, element):
				result = True
		return result
	else:
		return (to_lower_alpha(value) == to_lower_alpha(target))

def to_lower_alpha(string_in):
	return "".join(c for c in string_in if c in "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz").lower()

def format_date(in_date):
	return datetime.strftime(datetime.strptime(in_date.split(" ")[0], "%m/%d/%Y"), "%Y-%m-%d")

def clean_commodities_list(commodities_list: dict):
	#CONFIG.json contains an array of key:value pairs
	#loops through this array, matches value to commodities_list, and increments the corresponding key
	commodity_conversion_table = config.data["commodity_conversions"]

	#adds any missing keys to the commodity list
	for commodity in commodity_conversion_table:
		if not commodity in commodities_list:
			commodities_list[commodity] = 0

	#conversion loop
	for commodity in commodities_list:
		for key, value in commodity_conversion_table.items():
			if value == commodity:
				print(value + " converted to " + key)
				commodities_list[key] += commodities_list[value]

	#Kludge bug-fix: go through everything again and set all those converted SKUs' quantities to 0
	for commodity in commodities_list:
		for key, value in commodity_conversion_table.items():
			if value == commodity:
				print(value + " cleared")
				commodities_list[value] = 0

	return commodities_list

def create_output_folder(folder_string):
	folder_path = os.getcwd() + os.sep + folder_string
	if not os.path.exists(folder_path):
		os.mkdir(folder_path)

	error_file_file_path = folder_path + os.sep + config.data["default_date"] + os.sep + config.data["default_date"] + "-Error_file.txt"
	if os.path.exists(error_file_file_path):
		os.remove(error_file_file_path)

def clean_string(in_string):
	return "".join(c for c in in_string if c in "0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ ")

def info_box(input_string):
	app.infoBox("INFO", input_string)

def error_box(input_string):
	app.errorBox("ERROR", input_string)
	print(input_string)
	error_file(input_string)

def message_box(input_string):
	app.infoBox("INFO", input_string)
	print(input_string)

def error_file(input_string):
	try:
		with open(config.data["default_date"] + os.sep + config.data["default_date"] + "-Error_file.txt", "a") as file:
			file.write(input_string + "\n")
	except:
		error_box("Unable to write to error file!\nIs file open?")

def check_for_required_files_for_email():
	if app.getEntry("batchesFileEntry") == "":
		error_box("No Batches Scans file selected")
		return False
	if app.getEntry("ACEManifestFileEntry") == "":
		error_box("No ACE Manifest file selected")
		return False
	if app.getEntry("XLSXReportFileEntry") == "":
		error_box("No XLSX Report file selected")
		return False
	if app.getEntry("batchesFileEntry")[-4].lower() != ".csv":
		error_box("Batches Scans is not .csv")
		return False
	if app.getEntry("ACEManifestFileEntry")[-5].lower() != ".json":
		error_box("ACE Manifest is not .json")
		return False
	if app.getEntry("XLSXReportFileEntry")[-5].lower() != ".xlsx":
		error_box("XLSX Report is not .csv")
		return False
	return True

## API stuff

def get_token():
	if config.data["token"]["contents"] == None or config.data["token"]["creation_time"] == None:
		access_token = generate_token()
		return access_token
	else:
		#Checks if token is expired
		creation_time = datetime.strptime(config.data["token"]["creation_time"], "%Y-%m-%d %H:%M:%S")
		token_duration = config.data["token"]["contents"]["expires_in"]
		time_delta = datetime_module.timedelta(seconds = token_duration)
		target_time = creation_time + time_delta
		now = datetime.now()

		if now > target_time:
			access_token = generate_token()
		return access_token

def generate_token():
	host_url = "https://secure-wms.com/AuthServer/api/Token"
	headers = {
		"Content-Type": "application/json; charset=utf-8",
		"Accept": "application/json",
		"Host": "secure-wms.com",
		"Accept-Language": "Content-Length",
		"Accept-Encoding": "gzip,deflate,sdch",
		"Authorization": "Basic " + config.data["auth_key"]
	}
	payload = json.dumps({
		"grant_type": "client_credentials",
		"tpl": config.data["tpl"],
		"user_login_id": config.data["user_login_id"]
	})

	responce = requests.request("POST", host_url, data = payload, headers = headers, timeout = 3.0)

	#print("STATUS:", responce.status_code)
	#print(responce.text)
	access_token = responce.json()

	config.data["token"]["contents"] = access_token
	config.data["token"]["creation_time"] = datetime.strftime(datetime.now(), "%Y-%m-%d %H:%M:%S")
	config.save()
	logger.debug("3PLC Access token renewed")

	return access_token

def get_orders_from_batch(scan):

	if scan.length == 8:
		return scan
	elif scans.length == 6:
		get_token()
		#API stuff

### Start
config = Config()
trip = Trip()
app = gui()

if __name__ == "__main__":
	print("US TRUCK MANAGER v2.3")
	init_UI()